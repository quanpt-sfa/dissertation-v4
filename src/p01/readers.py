# pyright: reportMissingTypeStubs=false, reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false
"""Streaming raw-source readers and file-signature inspection."""

from __future__ import annotations

import csv
import hashlib
import json
from collections.abc import Iterator, Mapping
from datetime import UTC, datetime
from pathlib import Path
from typing import cast

import openpyxl
import pyarrow.parquet as pq

from .models import SourceSpec

Row = dict[str, object]


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def inspect_file_signature(path: Path, spec: SourceSpec) -> dict[str, object]:
    stat = path.stat()
    with path.open("rb") as stream:
        prefix = stream.read(32)
        stream.seek(max(stat.st_size - 4096, 0))
        tail = stream.read()
    bom = (
        "utf-8-sig"
        if prefix.startswith(b"\xef\xbb\xbf")
        else "utf-16-le"
        if prefix.startswith(b"\xff\xfe")
        else "utf-16-be"
        if prefix.startswith(b"\xfe\xff")
        else None
    )
    newline = "CRLF" if b"\r\n" in tail else "LF" if b"\n" in tail else "NONE_OR_BINARY"
    sha256 = hash_file(path)
    return {
        "path_relative": spec.relative_path,
        "format": spec.format,
        "size_bytes": stat.st_size,
        "modified_at_utc": datetime.fromtimestamp(stat.st_mtime, tz=UTC).isoformat(),
        "sha256": sha256,
        "locked_sha256": spec.locked_sha256,
        "hash_matches_lock": sha256 == spec.locked_sha256,
        "prefix_hex": prefix.hex(),
        "bom": bom,
        "newline_style": newline,
    }


def read_header(path: Path, spec: SourceSpec) -> list[str]:
    if spec.format in {"csv", "tsv"}:
        delimiter = spec.delimiter or ("\t" if spec.format == "tsv" else ",")
        with path.open("r", encoding=spec.encoding, newline="") as stream:
            reader = csv.reader(stream, delimiter=delimiter)
            try:
                return [str(value) for value in next(reader)]
            except StopIteration:
                return []
    if spec.format == "parquet":
        return list(pq.ParquetFile(path).schema_arrow.names)
    if spec.format == "xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[spec.sheet_name] if spec.sheet_name else workbook[workbook.sheetnames[0]]
        header_row = spec.header_row or 1
        values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        columns = ["" if value is None else str(value).strip() for value in values]
        workbook.close()
        return columns
    iterator = iter_rows(path, spec)
    try:
        return list(next(iterator))
    except StopIteration:
        return []


def iter_rows(path: Path, spec: SourceSpec) -> Iterator[Row]:
    if spec.format in {"csv", "tsv"}:
        delimiter = spec.delimiter or ("\t" if spec.format == "tsv" else ",")
        with path.open("r", encoding=spec.encoding, newline="") as stream:
            reader = csv.DictReader(stream, delimiter=delimiter)
            for row in reader:
                yield {str(key): value for key, value in row.items() if key is not None}
        return
    if spec.format == "parquet":
        parquet = pq.ParquetFile(path)
        for batch in parquet.iter_batches(batch_size=65536):
            for row in batch.to_pylist():
                yield {str(key): value for key, value in cast(Mapping[object, object], row).items()}
        return
    if spec.format == "xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[spec.sheet_name] if spec.sheet_name else workbook[workbook.sheetnames[0]]
        header_row = spec.header_row or 1
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        columns = ["" if value is None else str(value).strip() for value in header_values]
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if all(value is None for value in values):
                continue
            yield {column: value for column, value in zip(columns, values, strict=False)}
        workbook.close()
        return
    if spec.format == "jsonl":
        with path.open("r", encoding=spec.encoding) as stream:
            for line_number, line in enumerate(stream, start=1):
                if not line.strip():
                    continue
                value: object = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"JSONL row {line_number} is not an object")
                yield {str(key): item for key, item in cast(dict[object, object], value).items()}
        return
    if spec.format == "json":
        raw: object = json.loads(path.read_text(encoding=spec.encoding))
        if not isinstance(raw, list):
            raise ValueError("JSON source must be an array of objects")
        for index, value in enumerate(cast(list[object], raw), start=1):
            if not isinstance(value, dict):
                raise ValueError(f"JSON row {index} is not an object")
            yield {str(key): item for key, item in cast(dict[object, object], value).items()}
        return
    raise ValueError(f"Unsupported source format: {spec.format}")
