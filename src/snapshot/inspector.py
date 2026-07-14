# pyright: reportMissingTypeStubs=false, reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false
"""Deterministic schema and row-count inspection for snapshot discovery."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import openpyxl
import pyarrow.parquet as pq

from .models import ReaderSpec, SourceFormat


@dataclass(frozen=True)
class FileInspection:
    columns: tuple[str, ...]
    row_count: int
    dtypes: dict[str, str]
    reader: dict[str, object]
    schema_hash: str


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_name(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", ascii_value.casefold())


def resolve_semantics(
    columns: tuple[str, ...], candidates: dict[str, tuple[str, ...]]
) -> dict[str, str]:
    normalized: dict[str, list[str]] = {}
    for column in columns:
        normalized.setdefault(normalize_name(column), []).append(column)
    resolved: dict[str, str] = {}
    for semantic, options in candidates.items():
        matches: list[str] = []
        for candidate in options:
            matches.extend(normalized.get(normalize_name(candidate), []))
        unique = list(dict.fromkeys(matches))
        if len(unique) > 1:
            raise ValueError(f"semantic={semantic}: ambiguous candidate matches {unique}")
        if unique:
            resolved[semantic] = unique[0]
    return resolved


def inspect_file(path: Path, format_name: SourceFormat, reader: ReaderSpec) -> FileInspection:
    if format_name == "parquet":
        parquet = pq.ParquetFile(path)
        schema = parquet.schema_arrow
        columns = tuple(schema.names)
        row_count = parquet.metadata.num_rows if parquet.metadata is not None else 0
        dtypes = {field.name: str(field.type) for field in schema}
        reader_meta: dict[str, object] = {}
    elif format_name in {"csv", "tsv"}:
        delimiter = reader.delimiter or ("\t" if format_name == "tsv" else ",")
        with path.open("r", encoding=reader.encoding, newline="") as stream:
            csv_reader = csv.reader(stream, delimiter=delimiter)
            try:
                columns = tuple(str(value).strip() for value in next(csv_reader))
            except StopIteration:
                columns = ()
                row_count = 0
            else:
                row_count = sum(1 for _ in csv_reader)
        dtypes = {column: "string" for column in columns}
        reader_meta = {"encoding": reader.encoding, "delimiter": delimiter}
    elif format_name == "xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if reader.sheet_name in {None, "first"}:
            sheet = workbook[workbook.sheetnames[0]]
            sheet_name = workbook.sheetnames[0]
        else:
            if reader.sheet_name not in workbook.sheetnames:
                raise ValueError(f"file={path}: sheet {reader.sheet_name} not found")
            sheet_name = reader.sheet_name
            sheet = workbook[sheet_name]
        header_row = _resolve_header_row(sheet, reader.header_row)
        values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        columns = tuple("" if value is None else str(value).strip() for value in values)
        if not columns or any(not column for column in columns):
            raise ValueError(f"file={path}: blank or missing header cells at row {header_row}")
        row_count = max(sheet.max_row - header_row, 0)
        dtypes = {column: "dynamic_excel" for column in columns}
        reader_meta = {"sheet_name": sheet_name, "header_row": header_row}
        workbook.close()
    elif format_name == "jsonl":
        columns_set: set[str] = set()
        row_count = 0
        with path.open("r", encoding=reader.encoding) as stream:
            for line in stream:
                if not line.strip():
                    continue
                value: object = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"file={path}: JSONL rows must be objects")
                columns_set.update(str(key) for key in cast(dict[object, object], value))
                row_count += 1
        columns = tuple(sorted(columns_set))
        dtypes = {column: "dynamic_json" for column in columns}
        reader_meta = {"encoding": reader.encoding}
    elif format_name == "json":
        raw: object = json.loads(path.read_text(encoding=reader.encoding))
        if not isinstance(raw, list):
            raise ValueError(f"file={path}: JSON source must be an array")
        columns_set: set[str] = set()
        for value in cast(list[object], raw):
            if not isinstance(value, dict):
                raise ValueError(f"file={path}: JSON rows must be objects")
            columns_set.update(str(key) for key in cast(dict[object, object], value))
        columns = tuple(sorted(columns_set))
        row_count = len(raw)
        dtypes = {column: "dynamic_json" for column in columns}
        reader_meta = {"encoding": reader.encoding}
    else:
        raise ValueError(f"unsupported format {format_name}")

    payload = json.dumps(
        {"columns": columns, "dtypes": dtypes, "reader": reader_meta},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return FileInspection(
        columns=columns,
        row_count=row_count,
        dtypes=dtypes,
        reader=reader_meta,
        schema_hash=hashlib.sha256(payload).hexdigest(),
    )


def _resolve_header_row(sheet: Any, configured: int | str | None) -> int:
    if isinstance(configured, int):
        return configured
    if configured not in {None, "auto"}:
        raise ValueError("header_row must be integer or auto")
    best_row = 1
    best_count = -1
    # openpyxl worksheet is intentionally kept at the dynamic package boundary.
    for row_index, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=20, values_only=True),
        start=1,
    ):
        count = sum(1 for value in values if value is not None and str(value).strip())
        if count > best_count:
            best_count = int(count)
            best_row = row_index
    return best_row
